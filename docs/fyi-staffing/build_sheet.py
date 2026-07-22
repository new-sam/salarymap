# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

NAVY = "1F2A44"; ORANGE = "FB923C"; GREEN = "DCFCE7"; YELLOW = "FEF9C3"; RED = "FEE2E2"; BLUE = "E0F2FE"
FONT = "Malgun Gothic"
white = Font(color="FFFFFF", bold=True, size=11, name=FONT)
hdr_fill = PatternFill("solid", fgColor=NAVY)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

def f(size=10, bold=False, color="111111"):
    return Font(size=size, bold=bold, color=color, name=FONT)

# ===================== TAB 1 : 전략 요약 =====================
ws = wb.active
ws.title = "① 전략 요약"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 26
for c in ['C','D','E','F','G','H']:
    ws.column_dimensions[c].width = 20

def sec(row, text):
    for col in range(2, 9):
        cc = ws.cell(row=row, column=col)
        cc.fill = PatternFill("solid", fgColor=ORANGE)
    ws.cell(row=row, column=2, value=text).font = white
    ws.row_dimensions[row].height = 22

def kv(row, k, v, height=None):
    ws.cell(row=row, column=2, value=k).font = f(10, True, "374151")
    cc = ws.cell(row=row, column=3, value=v)
    cc.font = f(10); cc.alignment = wrap
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
    if height: ws.row_dimensions[row].height = height

ws.cell(2, 2, "FYI 베트남 스태핑 — 콜드 아웃리치 작업 시트").font = Font(bold=True, size=16, color=NAVY, name=FONT)
ws.cell(3, 2, "비용압박 한국 중소·스타트업  →  인건비 70%↓ 베트남 인재(개발·마케팅·디자인·운영) 리모트 매칭").font = f(11, False, "374151")
ws.merge_cells("C3:H3")

r = 5
sec(r, "이 시트 쓰는 법 — 3단계"); r += 1
for k, v in [("1. 담기", "공고 보고 적합한 회사를 해당 직무군 탭(②~⑤)에 행으로 추가"),
             ("2. 버킷 정하기", "아래 규칙으로 1/2/3순위 또는 [버림] 표시"),
             ("3. 발송·기록", "1순위부터 콜드 발송 → D+0/D+1/D+3 체크, 회신·미팅·LOI 기록")]:
    ws.cell(r, 2, k).font = f(11, True, NAVY)
    ws.cell(r, 3, v).font = f(11)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8); ws.row_dimensions[r].height = 20; r += 1
r += 1

sec(r, "버킷 규칙 — 한눈에"); r += 1
ws.cell(r, 2, "적합도 게이트").font = f(10, True)
ws.cell(r, 3, "공고 직무가 개발·마케팅·디자인·운영 중 하나? → 아니면 [버림]").font = f(10)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8); r += 1
for k, v, col in [("1순위 (즉시 발송)", "적합  +  채용 급함 (마감 후 재공고 / 30일+ 미충원)", GREEN),
                  ("2순위 (관망)", "적합  +  안 급함", YELLOW),
                  ("3순위 (테스트)", "직무 애매하지만 시도해볼 만", BLUE)]:
    cc = ws.cell(r, 2, k); cc.fill = PatternFill("solid", fgColor=col); cc.font = f(10, True); cc.border = border
    ws.cell(r, 3, v).font = f(10)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8); r += 1
r += 1

sec(r, "메시지 — 검증된 후크만"); r += 1
cc = ws.cell(r, 2, "✅ 쓸 것"); cc.fill = PatternFill("solid", fgColor=GREEN); cc.font = f(10, True); cc.border = border
ws.cell(r, 3, "\"베트남 시장 — 비용↓ 성과↑\"   ·   \"인건비 1/3로 줄이는 솔루션\"").font = f(10)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8); r += 1
cc = ws.cell(r, 2, "❌ 버릴 것"); cc.fill = PatternFill("solid", fgColor=RED); cc.font = f(10, True); cc.border = border
ws.cell(r, 3, "\"상위5% 월99만원\"   ·   멋사식 무차별 대량발송 (지난번 둘 다 회신 ~0%)").font = f(10)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8); r += 1
r += 1

sec(r, "세그먼트 태그 — 학습용"); r += 1
ws.cell(r, 2, "한 줄").font = f(10, True)
ws.cell(r, 3, "각 행에 베트남/소비재/일반 태그 → ⑥탭에서 어느 쪽이 잘 되는지 비교. 한 군에 몰빵 금지, 골고루 섞어 다음 라운드가 정답 정하게.").font = f(10)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8); ws.row_dimensions[r].height = 28; ws.cell(r, 3).alignment = wrap; r += 1

# ===================== Listing tabs =====================
COLS = ["No.","버킷","회사명","업종","기업구분","직원수","설립일","매출액","주소",
        "채용 포지션","담당자","직위","이메일","전화","채용링크",
        "세그먼트태그","베트남신호","채용소외신호","적합도점수","시급성점수",
        "메시지(후크)","D+0 발송","D+1","D+3","D+4 종료","회신","회신온도","미팅","LOI","비고"]
WIDTHS = [4,12,18,16,10,8,11,11,24, 18,11,9,24,13,16, 12,9,14,9,9, 16,9,7,7,9,7,10,7,6,22]

def make_listing(tab_name, rows):
    s = wb.create_sheet(tab_name)
    s.sheet_view.showGridLines = False
    for i, c in enumerate(COLS, start=1):
        cell = s.cell(1, i, c); cell.font = white; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
        s.column_dimensions[get_column_letter(i)].width = WIDTHS[i-1]
    s.freeze_panes = "C2"; s.row_dimensions[1].height = 30
    dv_bucket = DataValidation(type="list", formula1='"1순위,2순위,3순위,버림"', allow_blank=True)
    dv_seg = DataValidation(type="list", formula1='"베트남,소비재,일반"', allow_blank=True)
    dv_ox = DataValidation(type="list", formula1='"O,X,열람"', allow_blank=True)
    dv_temp = DataValidation(type="list", formula1='"Cold,Warm,Hot,LOI"', allow_blank=True)
    for dv in (dv_bucket, dv_seg, dv_ox, dv_temp): s.add_data_validation(dv)
    dv_bucket.add("B2:B500"); dv_seg.add("P2:P500")
    for col in ["V","W","X","Y","Z","AB","AC"]: dv_ox.add(col + "2:" + col + "500")
    dv_temp.add("AA2:AA500")
    rownum = 2
    for d in rows:
        d = list(d) + [""] * (len(COLS) - len(d))
        d[0] = str(rownum - 1)
        for i, val in enumerate(d, start=1):
            cell = s.cell(rownum, i, val); cell.font = f(10); cell.alignment = wrap; cell.border = border
        bcol = {"1순위":GREEN, "2순위":YELLOW, "3순위":BLUE, "버림":RED}.get(d[1])
        if bcol: s.cell(rownum, 2).fill = PatternFill("solid", fgColor=bcol)
        rownum += 1
    s.cell(rownum + 1, 3, "↑ 위는 지난 캠페인 실데이터 재배치(아카이브+재점수). 신규 발굴 행은 아래에 추가.").font = Font(italic=True, size=9, color="6B7280", name=FONT)
    return s

DEV = [
 ["1","1순위","에스더블유솔루션","솔루션·SI","중소기업","3","2025.01.20","5억","서울 송파구 법원로6길 7","개발(주니어2)","이왕우","대표","sw25@kakao.com","","","일반","N","","5","2","인건비 대비 유능한 인재","O","O","","","O","LOI","O","O","✅ LOI 성사 — 주니어 개발 2명"],
 ["2","1순위","지노시스","개발/AI","중소기업","","","","","개발(주니어3)","-","대표","","","","베트남","Y","","5","2","직접 전화(기존 KTC)","O","","","","O","LOI","O","O","✅ 기존 KTC·LOI — 덱 성공사례(2025 3명→2026 5명)"],
 ["3","2순위","코아아이티","컴퓨터·하드웨어","중소기업","131","2012.04.02","137억","충북 청주시 오창읍","개발","신동용","대표","support@coreit.co.kr","","https://www.jobkorea.co.kr/Recruit/GI_Read/48429672","일반","N","","4","1","인건비 대비 유능한 인재","O","O","","","X","","","","무응답 → 재시도 후보"],
 ["4","3순위","오픈마루","솔루션·SI·CRM","중소기업","22","2013.10.22","55억","대전 유성구 테크노3로 65","SW엔지니어","전준식","대표","sales@openmaru.io","","https://www.jobkorea.co.kr/Recruit/GI_Read/48341860","일반","N","","3","1","잡코리아 채용공고 후크","O","O","","","X","","","","무응답(채용공고형 후크=0%) → 메시지 교체 재시도"],
 ["5","3순위","오션정보기술","솔루션·SI","중소기업","101","2008","146억","대전 서구","초급/신입 개발","이종규","대표","oceaninsa@ocean-it.co.kr","","https://www.jobkorea.co.kr/Recruit/GI_Read/48333274","일반","N","","3","1","잡코리아 채용공고 후크","O","","","","X","","","","무응답 → 메시지 교체 재시도"],
 ["6","2순위","원정대","마케팅/AI","스타트업","","","","서울 금천구 가산디지털1로 168","AI/AX 개발자","담당자","-","biz@wonjd.com","070-8872-4151","groupby/원정대","일반","N","","4","","","","","","","","","","","🆕 groupby 발굴 · 이메일 검증완료"],
 ["7","3순위","볼트에이아이","클린테크/AI","스타트업","","","","","풀스택 엔지니어","담당자","-","","","groupby/볼트에이아이","일반","N","","3","","","","","","","","","","","🆕 발굴 · 이메일 수집필요"],
]
MKT = [
 ["1","1순위","모먼츠컴퍼니","화장품/커머스","-","","","","","마케터(주니어1)","담당자","-","contact@momentsco.com","","","소비재","Y","","5","2","인건비 1/3 솔루션","O","","O","","O","LOI","O","O","✅ LOI — 화장품 마케터"],
 ["2","1순위","디에프코퍼레이션","패션","-","","","","","마케터(시니어1)","담당자","-","","","","소비재","Y","","5","2","베트남 시장","O","","","","O","LOI","O","O","✅ LOI — 패션 마케터"],
 ["3","2순위","더웨이브스","화장품","-","","","","","마케터(시니어1,영어)","담당자","-","","","","소비재","Y","","5","1","베트남 시장","O","","","","O","Warm","O","","🟡 Warm — 미팅 성사"],
 ["4","2순위","이노바이드","화장품/덴탈","스타트업","","","","","마케팅","국진혁","대표","kiyang@dentlink.app","","","소비재","Y","","4","1","인건비 1/3 솔루션","O","","","","O","Warm","O","","🟡 Warm"],
 ["5","3순위","메이코더스","커머스/물류","스타트업","","","","","마케팅","최새미","대표","saemi@maycoders.com","","","베트남","Y","","3","1","베트남 시장","O","","","","O","Cold","","","회신만 — 팔로업 필요"],
 ["6","3순위","비누랩스","마케팅/AI","스타트업","","","","","마케팅","담당자","-","ad@vinulabs.com","","","베트남","Y","","3","1","인건비 1/3 솔루션","O","","","","O","Cold","","","회신만 — 팔로업 필요"],
 ["7","2순위","콘스탄트(theconst)","헤어케어/소비재","스타트업","","","","","글로벌 크리에이터 마케터(틱톡샵)","담당자","-","","","theconst.recruit.roundhr.com","소비재","N","","4","","","","","","","","","","","🆕 발굴 · 소비재+글로벌마케팅 적합 · 이메일 수집필요"],
 ["8","3순위","티지소사이어티(운전선생)","모빌리티","스타트업","","","","","콘텐츠 마케터","담당자","-","","","groupby/startups/126","일반","N","","3","","","","","","","","","","","🆕 발굴 · 이메일 수집필요"],
 ["9","3순위","아르템","라이브커머스/식품","스타트업","","","","","Brand Lead","담당자","-","","","artem.kr","소비재","N","","3","","","","","","","","","","","🆕 발굴 · 이메일 수집필요"],
]
DESIGN = [
 ["1","1순위","아시아비엔씨","화장품","-","","","","","마케터+디자이너(시니어)","담당자","-","","","","소비재","Y","","5","2","인건비 1/3 솔루션","O","","","","O","LOI","O","O","✅ LOI — 디자이너 포함"],
 ["2","3순위","리빌더에이아이","게임/VR/AI","스타트업","","","","","서비스 디자이너","담당자","-","","","groupby/리빌더에이아이","일반","N","","3","","","","","","","","","","","🆕 발굴 · 이메일 수집필요"],
]
OPS = [
 ["1","3순위","뉴셀렉트","커머스","스타트업","","","","","운영/CS","담당자","-","hr@newselect.co.kr","","","소비재","N","","3","1","가격+프로모션","O","","","","O","Cold","","","회신만 — 운영직무 테스트 후보"],
 ["2","3순위","유니브","에듀테크","스타트업","","","","","고객경험(CX) 파트장","담당자","-","","","groupby/유니브","일반","N","","3","","","","","","","","","","","🆕 발굴 · 이메일 수집필요"],
]

# ---- ingest bulk-collected leads (sheet_ready.csv) into matching 직무군 tabs ----
import csv as _csv, os as _os
def _load_leads():
    b = {'개발': [], '마케팅': [], '디자인': [], '운영·CS': []}
    p = 'sheet_ready.csv'
    if not _os.path.exists(p):
        return b
    order = ['No.','버킷','회사명','업종','기업구분','직원수','설립일','매출액','주소','채용 포지션',
             '담당자','직위','이메일','전화','채용링크','세그먼트태그','베트남신호','채용소외신호',
             '적합도점수','시급성점수','메시지(후크)','D+0 발송','D+1','D+3','D+4 종료',
             '회신','회신온도','미팅','LOI','비고']
    for r in _csv.DictReader(open(p, encoding='utf-8-sig')):
        jik = r.get('직무군', '')
        if jik not in b:
            jik = '운영·CS' if '운영' in jik else None
        if jik not in b:
            continue
        b[jik].append([r.get(c, '') for c in order])
    return b

_leads = _load_leads()
DEV += _leads['개발']; MKT += _leads['마케팅']; DESIGN += _leads['디자인']; OPS += _leads['운영·CS']
print('leads ingested:', {k: len(v) for k, v in _leads.items()})

make_listing("② 개발", DEV)
make_listing("③ 마케팅", MKT)
make_listing("④ 디자인", DESIGN)
make_listing("⑤ 운영·CS", OPS)

# ===================== TAB 6 : 세그먼트 트래킹 =====================
ws6 = wb.create_sheet("⑥ 세그먼트 트래킹")
ws6.sheet_view.showGridLines = False
ws6.cell(1, 2, "세그먼트별 회신율 비교 (발송 후 자동 학습)").font = Font(bold=True, size=14, color=NAVY, name=FONT)
ws6.cell(2, 2, "※ 직무군 탭에 발송·회신을 기록하면 여기서 어느 조합이 진짜 전환되는지 비교. 한 라운드 끝나면 가장 약한 세그먼트/메시지를 끄고 강한 쪽 확대.").font = f(10, False, "6B7280")
ws6.merge_cells("B2:H2"); ws6.cell(2, 2).alignment = wrap
heads = ["세그먼트","발송","회신","회신율","Warm+","LOI","비고"]
for i, h in enumerate(heads, start=2):
    c = ws6.cell(4, i, h); c.font = white; c.fill = hdr_fill; c.border = border; c.alignment = center
segs = ["개발","마케팅","디자인","운영·CS","── 태그별 ──","베트남 태그","소비재 태그","일반"]
for j, s in enumerate(segs, start=5):
    ws6.cell(j, 2, s).font = f(10, True)
    ws6.cell(j, 2).border = border
    for col in range(3, 9): ws6.cell(j, col).border = border
for col, w in zip("BCDEFGH", [16,8,8,9,8,6,28]):
    ws6.column_dimensions[col].width = w

out = "C:/Users/slsvm/AppData/Local/Temp/claude/C--Users-slsvm-Desktop-salarymap/417bb65a-edf6-4b9b-9376-3bd6f8a02145/scratchpad/FYI_staffing_master.xlsx"
import time as _t
saved = out
try:
    wb.save(out)
except PermissionError:
    saved = out.replace('.xlsx', f'_{int(_t.time())%100000}.xlsx')
    wb.save(saved)
print("SAVED", saved)
print("sheets:", wb.sheetnames)
with open('last_saved.txt','w',encoding='utf-8') as _f:
    _f.write(saved)
